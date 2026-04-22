"""
Build an example portfolio for Exercise 6 (agent analyzer).

Date: April 1, 2025 (first trading day of April 2025)
Portfolio value: $500,000
  - $50,000 in money market
  - $450,000 in stocks, allocated by sector market-cap weights

Queries the Rice Data Portal for:
  1. Sector market-cap weights on 2025-04-01
  2. Closing prices for selected stocks on 2025-04-01 and April 2020-2024
  3. Ticker metadata (sector, industry, size)

Then computes integer share counts so that each sector's allocation
matches its market-cap weight as closely as possible, splits each
stock's shares into 1-6 randomized purchase lots across the six April
dates (2020-2025), and writes an Excel file with full lot detail.
"""

import requests
import pandas as pd
import numpy as np
import math
import random
from datetime import date
from dotenv import load_dotenv, dotenv_values
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
import os

load_dotenv()
ACCESS_TOKEN = os.getenv("RICE_ACCESS_TOKEN")
if not ACCESS_TOKEN:
    config = dotenv_values(os.path.join(os.path.expanduser("~"), "repos", "busi722", ".env"))
    ACCESS_TOKEN = config.get("RICE_ACCESS_TOKEN")

API_URL = "https://data-portal.rice-business.org/api/query"

PORTFOLIO_VALUE = 500_000
MONEY_MARKET = 50_000
STOCK_ALLOCATION = PORTFOLIO_VALUE - MONEY_MARKET
DATE = "2025-04-01"
HISTORICAL_YEARS = range(2020, 2025)

# Selected stocks: top by market cap in each sector
STOCKS_BY_SECTOR = {
    "Technology": ["AAPL", "MSFT", "NVDA", "AVGO", "ORCL", "CRM", "CSCO", "IBM"],
    "Financial Services": ["BRK.B", "JPM", "V", "MA", "BAC", "WFC", "AXP"],
    "Healthcare": ["LLY", "UNH", "JNJ", "ABBV", "ABT"],
    "Consumer Cyclical": ["AMZN", "TSLA", "HD", "MCD", "TJX"],
    "Communication Services": ["GOOGL", "META", "NFLX", "TMUS", "T"],
    "Industrials": ["GE", "RTX", "CAT", "UNP"],
    "Consumer Defensive": ["WMT", "COST", "PG"],
    "Energy": ["XOM", "CVX", "COP"],
    "Real Estate": ["PLD", "AMT", "WELL"],
    "Basic Materials": ["LIN", "SHW", "FCX"],
    "Utilities": ["NEE", "SO", "DUK"],
}


def query(sql):
    """Query the Rice Data Portal."""
    r = requests.post(
        API_URL,
        headers={
            "Authorization": f"Bearer {ACCESS_TOKEN}",
            "Content-Type": "application/json",
        },
        json={"query": sql},
        timeout=120,
    )
    resp = r.json()
    if "error" in resp:
        raise RuntimeError(resp["error"])
    return pd.DataFrame(resp["data"], columns=resp["columns"])


def get_sector_weights():
    """Get sector market-cap weights on the target date."""
    sql = f"""
    SELECT t.sector, SUM(d.marketcap) as total_mcap
    FROM daily d
    JOIN tickers t ON d.ticker = t.ticker
    WHERE d.date = '{DATE}'
      AND d.marketcap IS NOT NULL
      AND t.sector IS NOT NULL AND t.sector != ''
      AND t.isdelisted = 'N'
    GROUP BY t.sector
    ORDER BY total_mcap DESC
    """
    df = query(sql)
    df["total_mcap"] = df["total_mcap"].astype(float)
    total = df["total_mcap"].sum()
    df["weight"] = df["total_mcap"] / total
    return df.set_index("sector")["weight"]


def get_prices():
    """Get closing prices for all selected stocks on the target date."""
    all_tickers = [t for tickers in STOCKS_BY_SECTOR.values() for t in tickers]
    tickers_str = "','".join(all_tickers)
    sql = f"""
    SELECT a.ticker, a.close as price
    FROM sep a
    WHERE a.date = '{DATE}'
      AND a.ticker IN ('{tickers_str}')
    """
    df = query(sql)
    df["price"] = df["price"].astype(float)
    return df.set_index("ticker")["price"]


def get_historical_prices():
    """Get closing prices on the first trading day of April for 2020-2024."""
    all_tickers = [t for tickers in STOCKS_BY_SECTOR.values() for t in tickers]
    tickers_str = "','".join(all_tickers)
    all_rows = []
    for year in HISTORICAL_YEARS:
        sql = f"""
        SELECT a.ticker, a.date, a.close as price
        FROM sep a
        WHERE a.date::DATE >= '{year}-04-01' AND a.date::DATE <= '{year}-04-07'
          AND a.ticker IN ('{tickers_str}')
        ORDER BY a.date::DATE
        """
        df = query(sql)
        df["date"] = pd.to_datetime(df["date"])
        df["price"] = df["price"].astype(float)
        first = df.groupby("ticker")["date"].min().reset_index()
        df = df.merge(first, on=["ticker", "date"])
        all_rows.append(df)
    return pd.concat(all_rows, ignore_index=True)


def get_ticker_metadata():
    """Get sector, industry, and size for all selected stocks."""
    all_tickers = [t for tickers in STOCKS_BY_SECTOR.values() for t in tickers]
    tickers_str = "','".join(all_tickers)
    sql = f"""
    SELECT ticker, sector, industry, scalemarketcap
    FROM tickers
    WHERE ticker IN ('{tickers_str}')
    """
    df = query(sql)
    return df.set_index("ticker").to_dict("index")


def allocate_shares(sector_weights, prices):
    """
    Allocate integer shares to each stock so that sector allocations
    match target weights. Within each sector, dollars are split
    equally among the selected stocks.
    """
    holdings = []

    for sector, tickers in STOCKS_BY_SECTOR.items():
        weight = sector_weights.get(sector, 0)
        sector_dollars = STOCK_ALLOCATION * weight
        per_stock = sector_dollars / len(tickers)

        for ticker in tickers:
            price = prices[ticker]
            shares = math.floor(per_stock / price)
            value = shares * price
            holdings.append({
                "sector": sector,
                "ticker": ticker,
                "price": price,
                "shares": shares,
                "value": value,
            })

    return pd.DataFrame(holdings)


def main():
    print(f"Portfolio date: {DATE}")
    print(f"Total portfolio value: ${PORTFOLIO_VALUE:,.0f}")
    print(f"Money market: ${MONEY_MARKET:,.0f}")
    print(f"Stock allocation: ${STOCK_ALLOCATION:,.0f}")
    print()

    # Get data
    print("Fetching sector weights...")
    sector_weights = get_sector_weights()
    print("Fetching stock prices...")
    prices = get_prices()
    print()

    # Sector weight table
    print("=" * 60)
    print("SECTOR MARKET-CAP WEIGHTS (April 1, 2025)")
    print("=" * 60)
    for sector in sector_weights.index:
        w = sector_weights[sector]
        dollars = STOCK_ALLOCATION * w
        print(f"  {sector:25s} {w*100:6.2f}%   ${dollars:>10,.0f}")
    print(f"  {'Money Market':25s}            ${MONEY_MARKET:>10,}")
    print(f"  {'':25s}            ${PORTFOLIO_VALUE:>10,}")
    print()

    # Allocate
    holdings = allocate_shares(sector_weights, prices)

    # Display
    print("=" * 75)
    print("PORTFOLIO HOLDINGS")
    print("=" * 75)
    total_invested = 0
    for sector in STOCKS_BY_SECTOR:
        sdf = holdings[holdings["sector"] == sector]
        sector_val = sdf["value"].sum()
        target = STOCK_ALLOCATION * sector_weights.get(sector, 0)
        total_invested += sector_val
        print(f"\n  {sector} (target ${target:,.0f}, actual ${sector_val:,.0f})")
        for _, row in sdf.iterrows():
            pct = row["value"] / PORTFOLIO_VALUE * 100
            print(
                f"    {row['ticker']:6s}  {int(row['shares']):>5,} shares"
                f"  @ ${row['price']:>8,.2f}  = ${row['value']:>10,.2f}"
                f"  ({pct:.2f}%)"
            )

    residual = STOCK_ALLOCATION - total_invested
    mm_total = MONEY_MARKET + residual
    print(f"\n  Stock total:  ${total_invested:>10,.2f}")
    print(f"  Residual cash added to money market: ${residual:>10,.2f}")
    print(f"  Money market: ${mm_total:>10,.2f}")
    print(f"  Grand total:  ${total_invested + mm_total:>10,.2f}")

    # Save holdings CSV
    output = holdings[["sector", "ticker", "shares", "price", "value"]].copy()
    output.loc[len(output)] = {
        "sector": "Money Market",
        "ticker": "CASH",
        "shares": 1,
        "price": mm_total,
        "value": mm_total,
    }
    output.to_csv("exercise6_portfolio.csv", index=False)
    print(f"\nSaved to exercise6_portfolio.csv")

    # --- Build purchase-lot Excel file ---
    print("\nFetching historical prices (2020-2024)...")
    hist = get_historical_prices()
    print("Fetching ticker metadata...")
    meta = get_ticker_metadata()

    # Build price lookup: ticker -> {date -> price}
    price_lookup = {}
    for _, row in hist.iterrows():
        price_lookup.setdefault(row["ticker"], {})[row["date"].date()] = row["price"]
    for _, row in holdings.iterrows():
        price_lookup.setdefault(row["ticker"], {})[date(2025, 4, 1)] = row["price"]
    all_dates = sorted(
        set(d for prices in price_lookup.values() for d in prices.keys())
    )

    # Split each stock's shares into 1-6 random lots
    random.seed(42)
    np.random.seed(42)
    all_lots = {}
    for sector, tickers in STOCKS_BY_SECTOR.items():
        for ticker in tickers:
            total_shares = int(
                holdings.loc[holdings["ticker"] == ticker, "shares"].iloc[0]
            )
            n_lots = random.randint(1, 6)
            lot_dates = sorted(random.sample(all_dates, n_lots))
            if n_lots == 1:
                shares_per_lot = [total_shares]
            else:
                cuts = sorted(
                    random.sample(
                        range(1, total_shares), min(n_lots - 1, total_shares - 1)
                    )
                )
                cuts = [0] + cuts + [total_shares]
                shares_per_lot = [cuts[i + 1] - cuts[i] for i in range(len(cuts) - 1)]
            lots = [
                {"date": d, "shares": s, "price": price_lookup[ticker][d]}
                for d, s in zip(lot_dates, shares_per_lot)
            ]
            all_lots[ticker] = {
                "sector": sector,
                "lots": lots,
                "total_shares": total_shares,
            }

    # Write Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Lots"

    header_font = Font(bold=True, size=11)
    sector_font = Font(bold=True, size=12, color="FFFFFF")
    sector_fill = PatternFill(
        start_color="2563EB", end_color="2563EB", fill_type="solid"
    )
    ticker_font = Font(bold=True, size=11)
    meta_font = Font(size=10, color="555555")
    thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

    row_num = 1
    ws.cell(row=row_num, column=1, value="Portfolio Purchase Lots").font = Font(
        bold=True, size=14
    )
    ws.cell(
        row=row_num + 1,
        column=1,
        value="Stocks purchased between April 2020 and April 2025",
    )
    row_num += 3

    for sector, tickers in STOCKS_BY_SECTOR.items():
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = sector_fill
            cell.font = sector_font
        ws.cell(row=row_num, column=1, value=sector)
        row_num += 1

        for ticker in tickers:
            info = all_lots[ticker]
            m = meta.get(ticker, {})
            industry = m.get("industry", "")
            size = m.get("scalemarketcap", "")

            ws.cell(row=row_num, column=1, value=ticker).font = ticker_font
            ws.cell(row=row_num, column=2, value=industry).font = meta_font
            ws.cell(row=row_num, column=3, value=size).font = meta_font
            row_num += 1

            ws.cell(row=row_num, column=2, value="Purchase Date").font = header_font
            ws.cell(row=row_num, column=3, value="Shares").font = header_font
            ws.cell(row=row_num, column=4, value="Price per Share").font = header_font
            ws.cell(row=row_num, column=5, value="Cost").font = header_font
            row_num += 1

            total_cost = 0
            for lot in info["lots"]:
                cost = lot["shares"] * lot["price"]
                total_cost += cost
                ws.cell(
                    row=row_num, column=2, value=lot["date"]
                ).number_format = "YYYY-MM-DD"
                ws.cell(row=row_num, column=3, value=lot["shares"])
                ws.cell(
                    row=row_num, column=4, value=lot["price"]
                ).number_format = "#,##0.00"
                ws.cell(
                    row=row_num, column=5, value=cost
                ).number_format = "#,##0.00"
                row_num += 1

            ws.cell(row=row_num, column=2, value="Total:").font = Font(italic=True)
            ws.cell(row=row_num, column=3, value=info["total_shares"]).font = Font(
                bold=True
            )
            ws.cell(row=row_num, column=5, value=total_cost).number_format = "#,##0.00"
            ws.cell(row=row_num, column=5).font = Font(bold=True)
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 2

        row_num += 1

    # Money market section
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = sector_fill
        cell.font = sector_font
    ws.cell(row=row_num, column=1, value="Money Market")
    row_num += 1
    ws.cell(row=row_num, column=1, value="CASH").font = ticker_font
    ws.cell(row=row_num, column=2, value="Balance:").font = header_font
    ws.cell(row=row_num, column=3, value=mm_total).number_format = "#,##0.00"
    ws.cell(row=row_num, column=3).font = Font(bold=True)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14

    xlsx_file = "exercise6_portfolio_lots.xlsx"
    wb.save(xlsx_file)
    print(f"Saved to {xlsx_file}")
    total_lots = sum(len(v["lots"]) for v in all_lots.values())
    print(f"  {len(all_lots)} stocks, {total_lots} total lots")


if __name__ == "__main__":
    main()
